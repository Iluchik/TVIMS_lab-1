from openpyxl import load_workbook
import plotly.graph_objects as go
import scipy.stats as sps
import numpy as np

nauka = load_workbook("./xlsx/nauka_1.xlsx")["2"]
demo = load_workbook("./xlsx/demo14.xlsx")["Возр. группы"]
indexes = load_workbook("./xlsx/index_tsen_tov.xlsx")["Данные"]

table_nauka = nauka["B5":"Q5"] + nauka["B7":"Q7"]
table_demo = (tuple(np.concatenate(demo["A13":"A20"])), tuple(np.concatenate(demo["AB13":"AB20"])))
table_indexes = indexes["D4":"AA4"] + indexes["D5":"AA5"]

def create_histogram(cells):
	data = {}
	for (cell_x, cell_y) in zip(cells[0], cells[1]):
		if isinstance(cell_x.value, str):
			data[cell_x.value.split()[0]] = round(cell_y.value, 1)
			continue
		data[cell_x.value] = round(cell_y.value, 1)
	fig = draw(data)
	return fig

def draw(list_data):
	fig = go.Figure()
	fig.add_trace(go.Scatter(x=list(list_data.keys()), y=list(list_data.values())))
	fig.add_hrect(y0=np.median(list(list_data.values())) * 0.998, y1=np.median(list(list_data.values())) * 1.002, line_width=0, annotation_text=f"Медиана = {np.median(list(list_data.values()))}", annotation_position="top left", fillcolor="gold", opacity=0.5)
	fig.add_hrect(y0=np.mean(list(list_data.values())) * 0.998, y1=np.mean(list(list_data.values())) * 1.002, line_width=0, annotation_text=f"Ср.знач. = {round(np.mean(list(list_data.values())), 2)}", annotation_position="bottom right", fillcolor="blue", opacity=0.5)
	if sps.mode(list(list_data.values())).count != 1:
		fig.add_hrect(y0=sps.mode(list(list_data.values())).mode * 0.998, y1=sps.mode(list(list_data.values())).mode * 1.002, line_width=0, annotation_text=f"Мода = {sps.mode(list(list_data.values())).mode}", annotation_position="bottom right", fillcolor="red", opacity=0.5)
	else:
		fig.add_hrect(y0=max(list(list_data.values())) * 0.998, y1=max(list(list_data.values())) * 1.002, line_width=0, annotation_text=f"Мода = {max(list(list_data.values()))}", annotation_position="bottom right", fillcolor="red", opacity=0.5)
	return fig

fig_nauka = create_histogram(table_nauka)
fig_nauka.update_yaxes(range=[1000, 5000])
fig_nauka.update_layout(legend_orientation="h",
				  legend=dict(x=.5, xanchor="center"),
				  title="Число организаций, выполнявших научные исследования и разработки, по секторам деятельности по Российской Федерации",
				  xaxis_title="Годы",
				  yaxis_title="Число организаций",
				  margin=dict(l=0, r=0, t=30, b=0))
fig_nauka.update_traces(hoverinfo="x+y", hovertemplate="Год: %{x}<br>Число организаций: %{y}<extra></extra>")
fig_nauka.show()

fig_demo = create_histogram(table_demo)
fig_demo.update_layout(legend_orientation="h",
				  legend=dict(x=.5, xanchor="center"),
				  title="Распределение населения по возрастным группам",
				  xaxis_title="Возрастная группа",
				  yaxis_title="Численность",
				  margin=dict(l=0, r=0, t=30, b=0))
fig_demo.update_traces(hoverinfo="x+y", hovertemplate="Возрастная группа: от %{x} лет<br>Численность: %{y}<extra></extra>")
fig_demo.show()

fig_indexes = create_histogram(table_indexes)
fig_indexes.update_yaxes(range=[70, 150])
fig_indexes.update_layout(legend_orientation="h",
				  legend=dict(x=.5, xanchor="center"),
				  title="Индексы цен производителей по товарам и товарным группам до 2010 г (процент)",
				  xaxis_title="Время",
				  yaxis_title="Аpматуpа пpомышленная тpубопpоводная, задвижки, затвоpы и запоpные узлы к ней,шт",
				  margin=dict(l=0, r=0, t=30, b=0))
fig_indexes.update_traces(hoverinfo="x+y", hovertemplate="Дата: %{x}<br>Количество: %{y}<extra></extra>")
fig_indexes.show()